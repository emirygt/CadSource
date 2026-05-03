"""
categories.py — Kategori CRUD + Excel import/export endpoint'leri
"""
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Turkish-aware case normalization: İ→i, Ğ→ğ, Ş→ş, Ç→ç, Ü→ü, Ö→ö then lower()
_TR_UPPER = str.maketrans("İĞŞÇÜÖI", "iğşçüöı")

def _norm(s: str) -> str:
    return s.strip().translate(_TR_UPPER).lower()

from db import get_db
from middleware.tenant import get_current_tenant, apply_tenant_schema

router = APIRouter(prefix="/categories", tags=["categories"])


class CategoryCreate(BaseModel):
    name: str
    color: Optional[str] = "#6366f1"


class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    color: Optional[str] = None


def _ensure_categories_table(db: Session):
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS categories (
            id         SERIAL PRIMARY KEY,
            name       VARCHAR NOT NULL,
            parent_id  INTEGER REFERENCES categories(id) ON DELETE SET NULL,
            color      VARCHAR DEFAULT '#6366f1',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """))
    db.execute(text("""
        DO $$ BEGIN
            ALTER TABLE categories ADD COLUMN IF NOT EXISTS parent_id INTEGER REFERENCES categories(id) ON DELETE SET NULL;
        EXCEPTION WHEN others THEN NULL; END $$;
    """))
    db.commit()


def _get_or_create_category(name: str, parent_id: Optional[int], db: Session) -> int:
    """Turkish-safe case-insensitive name + parent_id eşleşmesi; yoksa INSERT."""
    name = name.strip()
    name_norm = _norm(name)

    if parent_id is None:
        rows = db.execute(
            text("SELECT id, name FROM categories WHERE parent_id IS NULL")
        ).fetchall()
    else:
        rows = db.execute(
            text("SELECT id, name FROM categories WHERE parent_id = :pid"),
            {"pid": parent_id},
        ).fetchall()

    for row in rows:
        if _norm(row.name) == name_norm:
            return row.id

    result = db.execute(
        text("INSERT INTO categories (name, parent_id) VALUES (:name, :pid) RETURNING id"),
        {"name": name, "pid": parent_id},
    ).fetchone()
    return result.id


@router.get("")
def list_categories(
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    apply_tenant_schema(tenant, db)
    _ensure_categories_table(db)
    apply_tenant_schema(tenant, db)  # commit sonrası connection değişebilir, search_path yenile
    rows = db.execute(text(
        "SELECT id, name, parent_id, color, created_at FROM categories ORDER BY id"
    )).fetchall()
    counts = {
        r[0]: r[1]
        for r in db.execute(text(
            "SELECT category_id, COUNT(*) FROM cad_files WHERE category_id IS NOT NULL GROUP BY category_id"
        )).fetchall()
    }
    return [
        {
            "id": r.id,
            "name": r.name,
            "parent_id": r.parent_id,
            "color": r.color,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "file_count": counts.get(r.id, 0),
        }
        for r in rows
    ]


@router.post("", status_code=201)
def create_category(
    body: CategoryCreate,
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    apply_tenant_schema(tenant, db)
    _ensure_categories_table(db)
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Kategori adı boş olamaz")

    existing = db.execute(
        text("SELECT id, name FROM categories WHERE parent_id IS NULL")
    ).fetchall()
    if any(_norm(row.name) == _norm(name) for row in existing):
        raise HTTPException(status_code=400, detail="Bu isimde kategori zaten var")
    result = db.execute(
        text("INSERT INTO categories (name, color) VALUES (:name, :color) RETURNING id, name, color, created_at"),
        {"name": name, "color": body.color or "#6366f1"},
    ).fetchone()
    db.commit()
    return {"id": result.id, "name": result.name, "parent_id": None, "color": result.color,
            "created_at": result.created_at.isoformat(), "file_count": 0}


@router.put("/{cat_id}")
def update_category(
    cat_id: int,
    body: CategoryUpdate,
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    apply_tenant_schema(tenant, db)
    existing = db.execute(
        text("SELECT id, name, color FROM categories WHERE id = :id"),
        {"id": cat_id},
    ).fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    new_name  = body.name.strip()  if body.name  else existing.name
    new_color = body.color.strip() if body.color else existing.color
    if not new_name:
        raise HTTPException(status_code=400, detail="Kategori adı boş olamaz")
    db.execute(
        text("UPDATE categories SET name = :name, color = :color WHERE id = :id"),
        {"name": new_name, "color": new_color, "id": cat_id},
    )
    db.commit()
    return {"id": cat_id, "name": new_name, "color": new_color}


@router.delete("/{cat_id}")
def delete_category(
    cat_id: int,
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    apply_tenant_schema(tenant, db)
    existing = db.execute(
        text("SELECT id FROM categories WHERE id = :id"), {"id": cat_id}
    ).fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    db.execute(
        text("UPDATE cad_files SET category_id = NULL WHERE category_id = :id"),
        {"id": cat_id},
    )
    db.execute(text("DELETE FROM categories WHERE id = :id"), {"id": cat_id})
    db.commit()
    return {"status": "deleted", "id": cat_id}


@router.patch("/files/{file_id}")
def set_file_category(
    file_id: int,
    body: dict,
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    apply_tenant_schema(tenant, db)
    cat_id = body.get("category_id")
    if cat_id is not None:
        exists = db.execute(
            text("SELECT id FROM categories WHERE id = :id"), {"id": cat_id}
        ).fetchone()
        if not exists:
            raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    db.execute(
        text("UPDATE cad_files SET category_id = :cat_id WHERE id = :file_id"),
        {"cat_id": cat_id, "file_id": file_id},
    )
    db.commit()
    return {"status": "updated", "file_id": file_id, "category_id": cat_id}


# ── Excel Template ────────────────────────────────────────────────────────────

@router.get("/template")
def download_template(
    _tenant: dict = Depends(get_current_tenant),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kategoriler"

    headers = ["category_1", "category_2", "category_3"]
    header_fill = PatternFill("solid", fgColor="0EA5E9")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 24

    sample_rows = [
        ("Giyim", "Üst Giyim", "Mont"),
        ("Giyim", "Üst Giyim", "Kazak"),
        ("Giyim", "Alt Giyim", "Pantolon"),
        ("Giyim", "Alt Giyim", ""),
        ("Elektronik", "", ""),
        ("Elektronik", "Bilgisayar", "Laptop"),
        ("Ev", "Dekorasyon", ""),
        ("Ev", "Dekorasyon", "Tablo"),
        ("İç Giyim", "", ""),
    ]
    for row_data in sample_rows:
        ws.append(row_data)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=kategori_template.xlsx"},
    )


# ── Excel Import ──────────────────────────────────────────────────────────────

@router.post("/import")
def import_categories(
    file: UploadFile = File(...),
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sadece .xlsx veya .xls dosyası yükleyin")

    apply_tenant_schema(tenant, db)
    _ensure_categories_table(db)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.file.read()), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Dosya boş")

    # header satırını bul (category_1 içeren ilk satır)
    header_row_idx = None
    col_map: dict[str, int] = {}
    for i, row in enumerate(rows):
        norm = [str(c).strip().lower() if c else "" for c in row]
        if "category_1" in norm:
            header_row_idx = i
            for j, h in enumerate(norm):
                if h in ("category_1", "category_2", "category_3"):
                    col_map[h] = j
            break

    if header_row_idx is None or "category_1" not in col_map:
        raise HTTPException(status_code=400, detail="Başlık satırı bulunamadı. 'category_1' kolonu zorunlu.")

    data_rows = rows[header_row_idx + 1:]

    total_rows = 0
    success_count = 0
    errors = []

    for idx, row in enumerate(data_rows, start=header_row_idx + 2):
        def cell(key: str) -> str:
            col = col_map.get(key)
            if col is None:
                return ""
            val = row[col] if col < len(row) else None
            return str(val).strip() if val is not None else ""

        c1 = cell("category_1")
        c2 = cell("category_2")
        c3 = cell("category_3")

        if not c1:
            continue  # boş satır, atla

        total_rows += 1

        if c3 and not c2:
            errors.append({"row": idx, "message": f"category_3 ('{c3}') varken category_2 boş olamaz"})
            continue

        try:
            id1 = _get_or_create_category(c1, None, db)
            if c2:
                id2 = _get_or_create_category(c2, id1, db)
                if c3:
                    _get_or_create_category(c3, id2, db)
            db.commit()
            success_count += 1
        except Exception as e:
            db.rollback()
            errors.append({"row": idx, "message": str(e)})

    return {
        "total_rows": total_rows,
        "success_count": success_count,
        "error_count": len(errors),
        "errors": errors,
    }
