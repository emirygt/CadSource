"""
reports.py - tenant-scoped Excel/CSV exports for operational reporting.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Iterable, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import bindparam, text
from sqlalchemy.orm import Session

from db import get_db
from middleware.tenant import apply_tenant_schema, get_current_tenant
from routes.search import _normalize_status
from services.job_service import ensure_job_tables, json_loads

router = APIRouter(prefix="/reports", tags=["reports"])


def _stamp() -> str:
    return datetime.utcnow().strftime("%Y%m%d-%H%M%S")


def _fmt_dt(value) -> str:
    return value.isoformat(sep=" ", timespec="seconds") if value else ""


def _bool_text(value) -> str:
    return "Var" if bool(value) else "Yok"


def _make_workbook(title: str, rows: Iterable[dict], sheets: Optional[dict[str, Iterable[dict]]] = None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    default = wb.active
    default.title = title[:31] or "Report"
    _write_sheet(default, list(rows))
    if sheets:
        for sheet_name, sheet_rows in sheets.items():
            ws = wb.create_sheet(sheet_name[:31])
            _write_sheet(ws, list(sheet_rows))
    return wb


def _write_sheet(ws, rows: list[dict]) -> None:
    headers = list(rows[0].keys()) if rows else ["Bilgi"]
    if not rows:
        rows = [{"Bilgi": "Kayit bulunamadi"}]

    header_fill = PatternFill("solid", fgColor="0EA5E9")
    header_font = Font(bold=True, color="FFFFFF")
    ws.freeze_panes = "A2"

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, min(ws.max_row, 300) + 1):
            max_len = max(max_len, len(str(ws.cell(row=row_idx, column=col_idx).value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def _excel_response(filename: str, wb: openpyxl.Workbook) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _csv_response(filename: str, rows: list[dict]) -> StreamingResponse:
    buf = io.StringIO()
    headers = list(rows[0].keys()) if rows else ["Bilgi"]
    writer = csv.DictWriter(buf, fieldnames=headers)
    writer.writeheader()
    if rows:
        writer.writerows(rows)
    else:
        writer.writerow({"Bilgi": "Kayit bulunamadi"})
    content = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    return StreamingResponse(
        content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _report_response(base: str, fmt: str, rows: list[dict], *, sheets: Optional[dict[str, Iterable[dict]]] = None) -> StreamingResponse:
    fmt = (fmt or "xlsx").lower()
    if fmt == "csv":
        return _csv_response(f"{base}-{_stamp()}.csv", rows)
    if fmt != "xlsx":
        raise HTTPException(status_code=400, detail="Desteklenen formatlar: xlsx, csv")
    wb = _make_workbook(base[:31], rows, sheets=sheets)
    return _excel_response(f"{base}-{_stamp()}.xlsx", wb)


def _file_filter_query(
    *,
    search: Optional[str],
    approved: Optional[bool],
    status: Optional[str],
    category_id: Optional[int],
    file_format: Optional[str],
    has_preview: Optional[bool],
    has_clip: Optional[bool],
    has_file_data: Optional[bool],
    entity_min: Optional[int],
    entity_max: Optional[int],
    layer_min: Optional[int],
    layer_max: Optional[int],
    bbox_width_min: Optional[float],
    bbox_width_max: Optional[float],
    bbox_height_min: Optional[float],
    bbox_height_max: Optional[float],
    aspect_min: Optional[float],
    aspect_max: Optional[float],
    indexed_from: Optional[str],
    indexed_to: Optional[str],
    layer: Optional[str],
    entity_type: Optional[str],
    duplicate_status: Optional[str],
    group_id: Optional[int],
    sort_by: str,
    sort_dir: str,
    limit: int,
) -> tuple:
    params: dict = {"limit": int(limit)}
    clauses = []
    if search:
        params["search"] = f"%{search}%"
        clauses.append("f.filename ILIKE :search")
    if approved is not None:
        params["approved"] = approved
        clauses.append("f.approved = :approved")
    if status is not None:
        params["status"] = _normalize_status(status)
        clauses.append("f.approval_status = :status")
    if category_id is not None:
        params["category_id"] = category_id
        clauses.append("f.category_id = :category_id")
    if file_format:
        formats = [x.strip().lower() for x in str(file_format).split(",") if x.strip()]
        if formats:
            params["formats"] = formats
            clauses.append("LOWER(f.file_format) IN :formats")
    if has_preview is not None:
        clauses.append("(f.jpg_preview IS NOT NULL OR f.svg_preview IS NOT NULL)" if has_preview else "(f.jpg_preview IS NULL AND f.svg_preview IS NULL)")
    if has_clip is not None:
        clauses.append("f.clip_vector IS NOT NULL" if has_clip else "f.clip_vector IS NULL")
    if has_file_data is not None:
        clauses.append("f.file_data IS NOT NULL" if has_file_data else "f.file_data IS NULL")
    for key, value, clause in [
        ("entity_min", entity_min, "f.entity_count >= :entity_min"),
        ("entity_max", entity_max, "f.entity_count <= :entity_max"),
        ("layer_min", layer_min, "f.layer_count >= :layer_min"),
        ("layer_max", layer_max, "f.layer_count <= :layer_max"),
        ("bbox_width_min", bbox_width_min, "f.bbox_width >= :bbox_width_min"),
        ("bbox_width_max", bbox_width_max, "f.bbox_width <= :bbox_width_max"),
        ("bbox_height_min", bbox_height_min, "f.bbox_height >= :bbox_height_min"),
        ("bbox_height_max", bbox_height_max, "f.bbox_height <= :bbox_height_max"),
        ("aspect_min", aspect_min, "(f.bbox_width / NULLIF(f.bbox_height, 0)) >= :aspect_min"),
        ("aspect_max", aspect_max, "(f.bbox_width / NULLIF(f.bbox_height, 0)) <= :aspect_max"),
    ]:
        if value is not None:
            params[key] = value
            clauses.append(clause)
    if indexed_from:
        params["indexed_from"] = indexed_from
        clauses.append("f.indexed_at >= CAST(:indexed_from AS timestamp)")
    if indexed_to:
        params["indexed_to"] = indexed_to
        clauses.append("f.indexed_at <= (CAST(:indexed_to AS timestamp) + INTERVAL '1 day')")
    if layer:
        params["layer"] = layer
        clauses.append("f.layers::jsonb ? :layer")
    if entity_type:
        params["entity_type"] = entity_type.upper()
        clauses.append("f.entity_types::jsonb ? :entity_type")
    if duplicate_status:
        params["duplicate_status"] = duplicate_status
        clauses.append("f.duplicate_status = :duplicate_status")
    if group_id is not None:
        params["group_id"] = group_id
        clauses.append("""(f.duplicate_group_id = :group_id OR EXISTS (
            SELECT 1 FROM cad_file_group_members gm
            WHERE gm.file_id = f.id AND gm.group_id = :group_id
        ))""")

    sort_columns = {
        "indexed_at": "f.indexed_at",
        "filename": "f.filename",
        "entity_count": "f.entity_count",
        "layer_count": "f.layer_count",
        "bbox_width": "f.bbox_width",
        "bbox_height": "f.bbox_height",
        "file_format": "f.file_format",
        "duplicate_status": "f.duplicate_status",
    }
    sort_col = sort_columns.get(sort_by, "f.indexed_at")
    sort_direction = "ASC" if str(sort_dir).lower() == "asc" else "DESC"
    where = " WHERE " + " AND ".join(clauses) if clauses else ""
    stmt = text(f"""
        SELECT f.id, f.filename, f.filepath, f.file_format, f.entity_count, f.layer_count,
               f.bbox_width, f.bbox_height, f.bbox_area, f.indexed_at,
               f.approval_status, f.approved, f.approved_at,
               (f.jpg_preview IS NOT NULL OR f.svg_preview IS NOT NULL) AS has_preview,
               (f.clip_vector IS NOT NULL) AS has_clip,
               (f.file_data IS NOT NULL) AS has_file_data,
               f.content_hash, f.geometry_hash, f.duplicate_status, f.duplicate_group_id,
               c.name AS category_name
        FROM cad_files f
        LEFT JOIN categories c ON c.id = f.category_id
        {where}
        ORDER BY {sort_col} {sort_direction}, f.id DESC
        LIMIT :limit
    """)
    if "formats" in params:
        stmt = stmt.bindparams(bindparam("formats", expanding=True))
    return stmt, params


@router.get("/files")
def export_files_report(
    fmt: str = Query(default="xlsx", pattern="^(xlsx|csv)$"),
    limit: int = Query(default=10000, ge=1, le=50000),
    search: Optional[str] = Query(default=None),
    approved: Optional[bool] = Query(default=None),
    status: Optional[str] = Query(default=None),
    category_id: Optional[int] = Query(default=None),
    file_format: Optional[str] = Query(default=None),
    has_preview: Optional[bool] = Query(default=None),
    has_clip: Optional[bool] = Query(default=None),
    has_file_data: Optional[bool] = Query(default=None),
    entity_min: Optional[int] = Query(default=None, ge=0),
    entity_max: Optional[int] = Query(default=None, ge=0),
    layer_min: Optional[int] = Query(default=None, ge=0),
    layer_max: Optional[int] = Query(default=None, ge=0),
    bbox_width_min: Optional[float] = Query(default=None),
    bbox_width_max: Optional[float] = Query(default=None),
    bbox_height_min: Optional[float] = Query(default=None),
    bbox_height_max: Optional[float] = Query(default=None),
    aspect_min: Optional[float] = Query(default=None),
    aspect_max: Optional[float] = Query(default=None),
    indexed_from: Optional[str] = Query(default=None),
    indexed_to: Optional[str] = Query(default=None),
    layer: Optional[str] = Query(default=None),
    entity_type: Optional[str] = Query(default=None),
    duplicate_status: Optional[str] = Query(default=None),
    group_id: Optional[int] = Query(default=None),
    sort_by: str = Query(default="indexed_at"),
    sort_dir: str = Query(default="desc"),
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    apply_tenant_schema(tenant, db)
    stmt, params = _file_filter_query(
        search=search, approved=approved, status=status, category_id=category_id,
        file_format=file_format, has_preview=has_preview, has_clip=has_clip,
        has_file_data=has_file_data, entity_min=entity_min, entity_max=entity_max,
        layer_min=layer_min, layer_max=layer_max, bbox_width_min=bbox_width_min,
        bbox_width_max=bbox_width_max, bbox_height_min=bbox_height_min,
        bbox_height_max=bbox_height_max, aspect_min=aspect_min, aspect_max=aspect_max,
        indexed_from=indexed_from, indexed_to=indexed_to, layer=layer,
        entity_type=entity_type, duplicate_status=duplicate_status, group_id=group_id,
        sort_by=sort_by, sort_dir=sort_dir, limit=limit,
    )
    rows = db.execute(stmt, params).fetchall()
    report_rows = [
        {
            "ID": r.id,
            "Dosya": r.filename,
            "Format": (r.file_format or "").upper(),
            "Kategori": r.category_name or "",
            "Durum": r.approval_status or ("approved" if r.approved else "uploaded"),
            "Entity": int(r.entity_count or 0),
            "Katman": int(r.layer_count or 0),
            "BBox Genislik": float(r.bbox_width or 0),
            "BBox Yukseklik": float(r.bbox_height or 0),
            "BBox Alan": float(r.bbox_area or 0),
            "Preview": _bool_text(r.has_preview),
            "CLIP": _bool_text(r.has_clip),
            "Ham Dosya": _bool_text(r.has_file_data),
            "Duplicate Durumu": r.duplicate_status or "unique",
            "Duplicate Grup": r.duplicate_group_id or "",
            "Content Hash": r.content_hash or "",
            "Geometry Hash": r.geometry_hash or "",
            "Eklenme": _fmt_dt(r.indexed_at),
            "Dosya Yolu": r.filepath or "",
        }
        for r in rows
    ]
    return _report_response("cad-files-report", fmt, report_rows)


@router.get("/duplicates")
def export_duplicates_report(
    fmt: str = Query(default="xlsx", pattern="^(xlsx|csv)$"),
    group_type: Optional[str] = Query(default=None),
    duplicate_status: Optional[str] = Query(default=None),
    limit: int = Query(default=10000, ge=1, le=50000),
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    apply_tenant_schema(tenant, db)
    clauses = []
    params = {"limit": limit}
    if group_type:
        clauses.append("g.group_type = :group_type")
        params["group_type"] = group_type
    if duplicate_status:
        clauses.append("f.duplicate_status = :duplicate_status")
        params["duplicate_status"] = duplicate_status
    where = " WHERE " + " AND ".join(clauses) if clauses else ""
    rows = db.execute(text(f"""
        SELECT g.id AS group_id, g.group_type, g.title, gm.role, gm.score, gm.reason,
               f.id AS file_id, f.filename, f.file_format, f.duplicate_status,
               f.content_hash, f.geometry_hash, f.indexed_at
        FROM cad_file_groups g
        JOIN cad_file_group_members gm ON gm.group_id = g.id
        JOIN cad_files f ON f.id = gm.file_id
        {where}
        ORDER BY g.id DESC, CASE gm.role WHEN 'original' THEN 0 WHEN 'duplicate' THEN 1 ELSE 2 END, f.id
        LIMIT :limit
    """), params).fetchall()
    report_rows = [
        {
            "Grup ID": r.group_id,
            "Grup Tipi": r.group_type,
            "Baslik": r.title or "",
            "Rol": r.role or "",
            "Skor": float(r.score or 0) if r.score is not None else "",
            "Sebep": r.reason or "",
            "Dosya ID": r.file_id,
            "Dosya": r.filename,
            "Format": (r.file_format or "").upper(),
            "Duplicate Durumu": r.duplicate_status or "",
            "Content Hash": r.content_hash or "",
            "Geometry Hash": r.geometry_hash or "",
            "Eklenme": _fmt_dt(r.indexed_at),
        }
        for r in rows
    ]
    return _report_response("cad-duplicates-report", fmt, report_rows)


@router.get("/jobs")
def export_jobs_report(
    fmt: str = Query(default="xlsx", pattern="^(xlsx|csv)$"),
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    created_from: Optional[str] = Query(default=None),
    created_to: Optional[str] = Query(default=None),
    limit: int = Query(default=5000, ge=1, le=50000),
    tenant: dict = Depends(get_current_tenant),
    db: Session = Depends(get_db),
):
    ensure_job_tables(db)
    clauses = ["j.schema_name = :schema_name"]
    params = {"schema_name": tenant["schema_name"], "limit": limit}
    if status:
        clauses.append("j.status = :status")
        params["status"] = status
    if type:
        clauses.append("j.type = :type")
        params["type"] = type
    if created_from:
        clauses.append("j.created_at >= CAST(:created_from AS timestamp)")
        params["created_from"] = created_from
    if created_to:
        clauses.append("j.created_at <= (CAST(:created_to AS timestamp) + INTERVAL '1 day')")
        params["created_to"] = created_to
    where = " AND ".join(clauses)
    jobs = db.execute(text(f"""
        SELECT j.*
        FROM public.jobs j
        WHERE {where}
        ORDER BY j.created_at DESC
        LIMIT :limit
    """), params).fetchall()
    job_ids = [int(r.id) for r in jobs]
    item_rows = []
    if job_ids:
        item_stmt = text("""
            SELECT ji.*
            FROM public.job_items ji
            WHERE ji.job_id IN :ids
            ORDER BY ji.job_id DESC, ji.item_index, ji.id
        """).bindparams(bindparam("ids", expanding=True))
        item_rows = db.execute(item_stmt, {"ids": job_ids}).fetchall()

    report_rows = [
        {
            "Job ID": r.id,
            "Tip": r.type,
            "Durum": r.status,
            "Oncelik": r.priority,
            "Toplam": r.total_items,
            "Islenen": r.processed_items,
            "Basarili": r.succeeded_items,
            "Hatali": r.failed_items,
            "Kullanici": r.user_email or "",
            "Payload": str(json_loads(r.payload)),
            "Sonuc": str(json_loads(r.result)),
            "Hata": r.error or "",
            "Olusturma": _fmt_dt(r.created_at),
            "Baslama": _fmt_dt(r.started_at),
            "Bitis": _fmt_dt(r.finished_at),
        }
        for r in jobs
    ]
    item_report_rows = [
        {
            "Job ID": r.job_id,
            "Sira": r.item_index,
            "Dosya ID": r.file_id or "",
            "Dosya": r.filename or "",
            "Aksiyon": r.action or "",
            "Durum": r.status,
            "Mesaj": r.message or "",
            "Sonuc": str(json_loads(r.result)),
            "Baslama": _fmt_dt(r.started_at),
            "Bitis": _fmt_dt(r.finished_at),
        }
        for r in item_rows
    ]
    sheets = {"Job Items": item_report_rows} if fmt == "xlsx" else None
    return _report_response("cad-jobs-report", fmt, report_rows, sheets=sheets)
